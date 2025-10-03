import openpyxl

total = sum(row[1] for row in ws2.iter_rows(min_row=2, values_only=True))
print(f"Total: {total}")

# Prompt for Excel filename and sheet name

while True:
    filename = input("Enter the Excel filename for TB or ledger (e.g., sample_tb.xlsx): ")
    sheetname = input("Enter the sheet name to analyze (e.g., TB): ")
    try:
        wb2 = openpyxl.load_workbook(filename)
        ws2 = wb2[sheetname]
        break
    except Exception as e:
        print(f"Error loading file or sheet: {e}")
        print("Please try again.\n")

print("Trial Balance:")
for row in ws2.iter_rows(min_row=2, values_only=True):
    print(f"Account: {row[0]}, Amount: {row[1]}")

total = sum(row[1] for row in ws2.iter_rows(min_row=2, values_only=True))
print(f"Total: {total}")
import random
def pick_samples(ws, n=2):
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    samples = random.sample(rows, min(n, len(rows)))
    print(f"\nRandomly selected {len(samples)} samples:")
    for account, amount in samples:
        print(f"Account: {account}, Amount: {amount}")

pick_samples(ws2, n=2)

# Flag unusual amounts (e.g., amounts above 10,000 or below -5,000)
def flag_anomalies(ws, high=10000, low=-5000):
    print("\nFlagged anomalies:")
    for account, amount in ws.iter_rows(min_row=2, values_only=True):
        if amount > high or amount < low:
            print(f"Account: {account}, Amount: {amount}")

flag_anomalies(ws2)
def flag_specific_accounts(ws, accounts):
    print("\nFlagged specific accounts:")
    for account, amount in ws.iter_rows(min_row=2, values_only=True):
        if account in accounts:
            print(f"Account: {account}, Amount: {amount}")

def flag_zscore_anomalies(ws, threshold=2):
    import statistics
    amounts = [row[1] for row in ws.iter_rows(min_row=2, values_only=True)]
    mean = statistics.mean(amounts)
    stdev = statistics.stdev(amounts)
    print("\nFlagged z-score anomalies:")
    for account, amount in ws.iter_rows(min_row=2, values_only=True):
        z = (amount - mean) / stdev if stdev else 0
        if abs(z) > threshold:
            print(f"Account: {account}, Amount: {amount}, Z-score: {z:.2f}")

def main_menu():
    print("\nChoose anomaly detection method:")
    print("1. Custom thresholds")
    print("2. Specific accounts")
    print("3. Statistical (z-score)")
    choice = input("Enter option (1/2/3): ")
    if choice == "1":
        high = float(input("Enter high threshold: "))
        low = float(input("Enter low threshold: "))
        flag_anomalies(ws2, high, low)
    elif choice == "2":
        accounts = input("Enter account names separated by comma: ").split(",")
        accounts = [a.strip() for a in accounts]
        flag_specific_accounts(ws2, accounts)
    elif choice == "3":
        threshold = float(input("Enter z-score threshold (e.g., 2): "))
        flag_zscore_anomalies(ws2, threshold)
    else:
        print("Invalid option.")

main_menu()

# --- AI Integration ---

# --- AI Integration using openai ---
import os
import openai

def get_openai_api_key():
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        print("OpenAI API key not set. Please set $env:OPENAI_API_KEY in your terminal.")
        return None
    return api_key

def openai_chat(prompt):
    api_key = get_openai_api_key()
    if not api_key:
        return None
    openai.api_key = api_key
    try:
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[{"role": "user", "content": prompt}]
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        print(f"OpenAI error: {e}")
        return None

def explain_anomalies(anomalies):
    if not anomalies:
        return
    prompt = "Explain the following flagged anomalies in an audit context:\n" + "\n".join(anomalies)
    response = openai_chat(prompt)
    print("\nAI Explanation:")
    print(response)

def generate_working_paper(anomalies):
    if not anomalies:
        return
    prompt = "Draft an audit working paper summary for these flagged anomalies:\n" + "\n".join(anomalies)
    response = openai_chat(prompt)
    print("\nAI-Generated Working Paper:")
    print(response)

# Example usage after anomaly detection:
if __name__ == "__main__":
    # Collect anomalies from custom threshold detection
    anomalies = []
    for account, amount in ws2.iter_rows(min_row=2, values_only=True):
        if amount > 10000 or amount < -5000:
            anomalies.append(f"Account: {account}, Amount: {amount}")
    if anomalies:
        explain_anomalies(anomalies)
        generate_working_paper(anomalies)
